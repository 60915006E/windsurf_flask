#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 匯出模組

使用 pandas 與 openpyxl 作為核心引擎
支援 Oracle LISTAGG 產生的換行資料處理

作者：系統管理員
建立日期：2026-03-17
版本：1.0.0
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional

# 設定日誌
logger = logging.getLogger(__name__)

class ExcelExporter:
    """Excel 匯出器類別"""
    
    def __init__(self):
        """初始化 Excel 匯出器"""
        self.workbook = None
        self.worksheet = None
        
    def clean_data(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        資料清洗邏輯
        
        處理 Oracle LISTAGG 產生的換行資料，將 <br> 或 <br/> 標籤替換為 Excel 可識別的換行符號
        
        Args:
            data: 原始資料列表
            
        Returns:
            List[Dict[str, Any]]: 清洗後的資料列表
        """
        try:
            cleaned_data = []
            
            for row in data:
                cleaned_row = {}
                for key, value in row.items():
                    if value is None:
                        cleaned_row[key] = ''
                    elif isinstance(value, str):
                        # 替換 HTML 換行標籤為 Excel 換行符號
                        cleaned_value = value.replace('<br>', '\n').replace('<br/>', '\n').replace('<br />', '\n')
                        # 移除其他 HTML 標籤（如果有）
                        import re
                        cleaned_value = re.sub(r'<[^>]+>', '', cleaned_value)
                        # 去除首尾空白
                        cleaned_value = cleaned_value.strip()
                        cleaned_row[key] = cleaned_value
                    else:
                        cleaned_row[key] = value
                
                cleaned_data.append(cleaned_row)
            
            logger.info(f"資料清洗完成，處理 {len(cleaned_data)} 筆資料")
            return cleaned_data
            
        except Exception as e:
            logger.error(f"資料清洗失敗: {str(e)}")
            raise RuntimeError(f"資料清洗失敗: {str(e)}")
    
    def apply_field_mapping(self, data: List[Dict[str, Any]], field_mapping: Dict[str, str]) -> List[Dict[str, Any]]:
        """
        動態欄位更名
        
        將資料庫欄位名稱轉換為對應的中文字稱
        
        Args:
            data: 原始資料列表
            field_mapping: 欄位映射字典 {field_id: display_name}
            
        Returns:
            List[Dict[str, Any]]: 欄位更名後的資料列表
        """
        try:
            mapped_data = []
            
            for row in data:
                mapped_row = {}
                for field_id, value in row.items():
                    # 取得顯示名稱，如果沒有映射則使用原始欄位名
                    display_name = field_mapping.get(field_id, field_id)
                    mapped_row[display_name] = value
                
                mapped_data.append(mapped_row)
            
            logger.info(f"欄位映射完成，映射 {len(field_mapping)} 個欄位")
            return mapped_data
            
        except Exception as e:
            logger.error(f"欄位映射失敗: {str(e)}")
            raise RuntimeError(f"欄位映射失敗: {str(e)}")
    
    def create_dataframe(self, data: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        建立 DataFrame
        
        Args:
            data: 處理後的資料列表
            
        Returns:
            pd.DataFrame: DataFrame 物件
        """
        try:
            if not data:
                logger.warning("沒有資料可以匯出")
                return pd.DataFrame()
            
            df = pd.DataFrame(data)
            logger.info(f"DataFrame 建立成功，形狀: {df.shape}")
            return df
            
        except Exception as e:
            logger.error(f"DataFrame 建立失敗: {str(e)}")
            raise RuntimeError(f"DataFrame 建立失敗: {str(e)}")
    
    def style_worksheet(self, df: pd.DataFrame) -> None:
        """
        美化工作表樣式
        
        Args:
            df: DataFrame 物件
        """
        try:
            # 設定標題樣式
            header_font = Font(name='Microsoft JhengHei', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 設定邊框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 套用標題樣式
            for cell in self.worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 設定資料欄位樣式
            data_font = Font(name='Microsoft JhengHei', size=10)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            for row in self.worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            # 調整欄位寬度
            for column in self.worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # 設定最小寬度和最大寬度
                adjusted_width = min(max(max_length + 2, 8), 50)
                self.worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info("工作表樣式設定完成")
            
        except Exception as e:
            logger.error(f"工作表樣式設定失敗: {str(e)}")
            # 樣式設定失敗不影響主要功能，只記錄錯誤
    
    def export_to_excel(self, data: List[Dict[str, Any]], field_mapping: Dict[str, str], 
                       sheet_name: str = "搜尋結果") -> BytesIO:
        """
        匯出資料到 Excel 檔案
        
        Args:
            data: 原始資料列表
            field_mapping: 欄位映射字典
            sheet_name: 工作表名稱
            
        Returns:
            BytesIO: Excel 檔案內容
        """
        try:
            logger.info(f"開始匯出 Excel，資料筆數: {len(data)}")
            
            # 1. 資料清洗
            cleaned_data = self.clean_data(data)
            
            # 2. 欄位映射
            mapped_data = self.apply_field_mapping(cleaned_data, field_mapping)
            
            # 3. 建立 DataFrame
            df = self.create_dataframe(mapped_data)
            
            if df.empty:
                logger.warning("沒有資料可以匯出")
                return BytesIO()
            
            # 4. 建立 Excel 工作簿
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = sheet_name
            
            # 5. 寫入資料到工作表
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    self.worksheet.cell(row=r_idx, column=c_idx, value=value)
            
            # 6. 美化工作表樣式
            self.style_worksheet(df)
            
            # 7. 儲存到 BytesIO
            excel_file = BytesIO()
            self.workbook.save(excel_file)
            excel_file.seek(0)
            
            logger.info(f"Excel 匯出成功，檔案大小: {len(excel_file.getvalue())} bytes")
            return excel_file
            
        except Exception as e:
            logger.error(f"Excel 匯出失敗: {str(e)}")
            raise RuntimeError(f"Excel 匯出失敗: {str(e)}")
        finally:
            # 清理資源
            if self.workbook:
                self.workbook.close()
                self.workbook = None
                self.worksheet = None

def export_search_results(data: List[Dict[str, Any]], field_mapping: Dict[str, str], 
                         filename_prefix: str = "搜尋結果") -> tuple:
    """
    匯出搜尋結果到 Excel
    
    Args:
        data: 搜尋結果資料
        field_mapping: 欄位映射字典
        filename_prefix: 檔名前綴
        
    Returns:
        tuple: (BytesIO 檔案內容, 檔案名稱)
    """
    try:
        # 生成檔名（包含當前時間）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        
        # 建立 Excel 匯出器
        exporter = ExcelExporter()
        
        # 匯出 Excel
        excel_file = exporter.export_to_excel(data, field_mapping)
        
        logger.info(f"搜尋結果 Excel 匯出完成: {filename}")
        return excel_file, filename
        
    except Exception as e:
        logger.error(f"搜尋結果 Excel 匯出失敗: {str(e)}")
        raise RuntimeError(f"搜尋結果 Excel 匯出失敗: {str(e)}")

# ===========================================
# 工具函數
# ===========================================

def check_dependencies():
    """
    檢查必要的依賴套件
    
    Returns:
        bool: 是否所有依賴都可用
    """
    try:
        import pandas
        import openpyxl
        logger.info(f"pandas 版本: {pandas.__version__}")
        logger.info(f"openpyxl 版本: {openpyxl.__version__}")
        return True
    except ImportError as e:
        logger.error(f"依賴套件檢查失敗: {str(e)}")
        return False

def get_field_mapping_from_settings(field_settings: List[Dict[str, Any]]) -> Dict[str, str]:
    """
    從欄位設定建立映射字典
    
    Args:
        field_settings: 欄位設定列表
        
    Returns:
        Dict[str, str]: 欄位映射字典
    """
    try:
        mapping = {}
        for field in field_settings:
            field_id = field.get('FIELD_ID', '')
            display_name = field.get('DISPLAY_NAME', '')
            if field_id and display_name:
                mapping[field_id] = display_name
        
        logger.info(f"欄位映射字典建立完成，包含 {len(mapping)} 個欄位")
        return mapping
        
    except Exception as e:
        logger.error(f"欄位映射字典建立失敗: {str(e)}")
        return {}

if __name__ == "__main__":
    # 測試模組
    print("Excel 匯出模組測試")
    
    # 檢查依賴
    if check_dependencies():
        print("✅ 所有依賴套件可用")
    else:
        print("❌ 依賴套件檢查失敗")
    
    # 測試資料
    test_data = [
        {
            'RP_NO': 'R001',
            'RP_NAME': '測試報告<br>包含換行',
            'MAIN_AUTHOR': '張三',
            'RP_KEYWORD': '關鍵字1<br/>關鍵字2<br />關鍵字3'
        },
        {
            'RP_NO': 'R002', 
            'RP_NAME': '另一個測試',
            'MAIN_AUTHOR': '李四',
            'RP_KEYWORD': '單一關鍵字'
        }
    ]
    
    test_field_mapping = {
        'RP_NO': '報告編號',
        'RP_NAME': '報告名稱',
        'MAIN_AUTHOR': '主要作者',
        'RP_KEYWORD': '關鍵字'
    }
    
    try:
        excel_file, filename = export_search_results(test_data, test_field_mapping)
        print(f"✅ 測試成功，檔案: {filename}, 大小: {len(excel_file.getvalue())} bytes")
    except Exception as e:
        print(f"❌ 測試失敗: {str(e)}")
